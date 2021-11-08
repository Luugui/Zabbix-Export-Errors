from pyzabbix import ZabbixAPI
from colorama import Fore, init
from openpyxl import Workbook
from openpyxl.styles import Font
from tqdm import tqdm
import argparse, os

init(autoreset=True)

os.system('cls') if os.name == 'nt' else os.system('clear')
print(f"""{Fore.GREEN}
  _____     _     _     _      _____                       _   _____
 |__  /__ _| |__ | |__ (_)_  _| ____|_  ___ __   ___  _ __| |_| ____|_ __ _ __ ___  _ __
   / // _` | '_ \| '_ \| \ \/ /  _| \ \/ / '_ \ / _ \| '__| __|  _| | '__| '__/ _ \| '__|
  / /| (_| | |_) | |_) | |>  <| |___ >  <| |_) | (_) | |  | |_| |___| |  | | | (_) | |
 /____\__,_|_.__/|_.__/|_/_/\_\_____/_/\_\ .__/ \___/|_|   \__|_____|_|  |_|  \___/|_|
                                         |_|
{Fore.WHITE}
 * Zabbix Export Error
 * Version: 1.1.0
 * Author: Luis Amaral
""")

parser = argparse.ArgumentParser(description="Extract trend from select host")
parser.add_argument("-u", "--user", required=True, help="Zabbix user")
parser.add_argument("-p", "--password", required=True, help="Zabbix user password")
parser.add_argument("-s", "--server", required=True, help="Zabbix frontend URL")

args = vars(parser.parse_args())

#  REPORT'S VARIABLES
URL=args['server']
LOGIN=args['user']
SENHA=args['password']


app = ZabbixAPI(URL)
try:
    if "https" in URL:
      import requests

      requests.packages.urllib3.disable_warnings()
      app.session.verify = False
    app.login(LOGIN,SENHA)
    print(f"{Fore.GREEN}+{Fore.WHITE} Connected")
except:
    print(f"{Fore.RED}-{Fore.WHITE} Not connected")

def get_items_id():
    items_id = []
    items = app.item.get(output=["itemid"],filter={"state": 1})
    items_id = [iid['itemid'] for iid in items]
    return items_id


def get_error_items(iid):
    items = {}
    error = app.item.get(output=["name","error"],selectHosts=["host"],itemids=iid)
    items = {"host": error[0]['hosts'][0]['host'],"item": error[0]['name'],"error": error[0]['error']}
    # if len(sys.argv) > 1 and sys.argv[1] == "-v":
    #     print(f"{Fore.GREEN}+{Fore.WHITE} Host={Fore.BLUE}{items['host']}{Fore.WHITE} || Name={Fore.BLUE}{items['item']}{Fore.WHITE} || Error={Fore.BLUE}{items['error']}{Fore.WHITE}")
    return items


def main():

    # Get the ids from items with error status
    list_ids = get_items_id()

    # Create the excel file and configure the header
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Items"
    sheet["A1"] = "Host"
    sheet["A1"].font = Font(sz=12, bold=True)
    sheet["B1"] = "Item"
    sheet["B1"].font = Font(sz=12, bold=True)
    sheet["C1"] = "Error"
    sheet["C1"].font = Font(sz=12, bold=True)

    row=2
    col_host=0
    col_item=0
    col_error=0

    for it in tqdm(list_ids,desc=f"{Fore.GREEN}+{Fore.WHITE} Generate report"):

        # Get the host, name and error message from a item
        error = get_error_items(it)

        # Input values in sheet
        sheet.cell(row=row, column=1).value = error['host']
        if len(error['host']) > col_host:
            col_host = len(error['host'])
        sheet.cell(row=row, column=2).value = error['item']
        if len(error['item']) > col_item:
            col_item = len(error['item'])
        sheet.cell(row=row, column=3).value = error['error']
        if len(error['error']) > col_error:
            col_error = len(error['error'])

        # Increment row variable
        row += 1


    # Excel configuration
    area = sheet.dimensions
    sheet.auto_filter.ref = area
    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = col_host + 3
    sheet.column_dimensions["B"].width = col_item
    sheet.column_dimensions["C"].width = col_error

    # Save Excel file report
    try:
        wb.save("Zabbix-Export-Items-Error.xlsx")
        print(f"{Fore.GREEN}+{Fore.WHITE} Report Save")
    except:
        print(f"{Fore.RED}-{Fore.WHITE} Error save report! Try again")
    # Logout from  User API
    app.user.logout()

    # for it in items_error:
    #     print(f"{Fore.GREEN}+{Fore.WHITE} Host={Fore.BLUE}{it['hosts'][0]['host']}{Fore.WHITE} || Name={Fore.BLUE}{it['name']}{Fore.WHITE} || Error={Fore.BLUE}{it['error']}{Fore.WHITE}")

main()
